#!/usr/bin/env python3
"""Convert salary data from Excel to JSON format.

Ported from MadsLorentzen/ai-job-search tools/convert_salary_excel.py (upstream fix 621ce5a).

Key fix: parse_numeric_cell now rejects ambiguous dot/comma thousands separators
(e.g. European "60.000" that means 60,000 but would parse as 60.0), preventing
a 1000x under-parsing bug on localized salary exports.

Prerequisites:
    pip install openpyxl

Usage:
    python tools/convert_salary_excel.py <path-to-excel-file>
    python tools/convert_salary_excel.py <path-to-excel-file> --source "Company Stats 2025"
    python tools/convert_salary_excel.py <path-to-excel-file> --baseline 100

Output: salary_data.json in the repo root.
"""

import json
import sys
import argparse
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Column name patterns for auto-detection
COMPANY_PATTERNS = {"company", "employer", "firm", "organization"}
CITY_PATTERNS = {"city", "location", "state", "region"}
COUNT_PATTERNS = {"count", "number", "n", "employees", "headcount", "antal"}
INDEX_PATTERNS = {"index", "idx", "salary", "median", "average", "compensation", "pay"}
COMPOUND_PATTERNS = {"antal", "indeks", "løn", "gennemsnit", "medarbejdere"}
ID_PATTERNS = {"id", "personnummer"}


def parse_numeric_cell(value):
    """Parse numeric Excel values, including localized string cells.

    Upstream fix 621ce5a: rejects ambiguous dot/comma thousands separators.
    E.g. European "60.000" means 60,000 but would silently parse as 60.0
    without this guard — a 1000x error on localized salary exports.
    """
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        raise ValueError("not numeric")

    text = value.strip().replace("\u00a0", " ").replace(" ", "")
    if not text:
        raise ValueError("not numeric")

    if "," in text and "." in text:
        # e.g. "1,234.56" (US) or "1.234,56" (EU) — strip thousands, normalise decimal
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        # e.g. "60,000" (US thousands) vs "60,5" (EU decimal)
        if re.fullmatch(r"[+-]?\d+,\d{3}", text):
            raise ValueError("ambiguous comma separator")
        text = text.replace(",", ".")
    elif "." in text:
        # e.g. "60.000" (EU thousands — ambiguous!) vs "60.5" (US decimal)
        if re.fullmatch(r"[+-]?\d+\.\d{3}", text):
            raise ValueError("ambiguous dot separator")

    return float(text)


def header_matches(header, patterns):
    h = header.lower().strip()
    tokens = set(re.findall(r"[a-z0-9]+", h))
    for p in patterns:
        if p in tokens:
            return True
        if p in COMPOUND_PATTERNS and p in h:
            return True
    return False


def detect_column_type(header):
    if header_matches(header, COUNT_PATTERNS):
        return "count"
    if header_matches(header, INDEX_PATTERNS):
        return "index"
    return None


def strip_type_patterns(header, patterns):
    name = header.lower()
    for p in patterns:
        name = re.sub(rf"(?<![a-z0-9]){re.escape(p)}(?![a-z0-9])", "", name)
    return name.strip(" _-")


def parse_sheet(ws, sheet_label=None):
    """Parse a single worksheet into a list of company salary entries."""
    # Find header row (scan first 10 rows)
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
        for cell in row:
            if cell.value and header_matches(str(cell.value), COMPANY_PATTERNS):
                header_row = row_idx
                break
        if header_row:
            break

    if header_row is None:
        print(f"Warning: Could not find header row in sheet '{ws.title}'. Skipping.", file=sys.stderr)
        return []

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[header_row]]

    company_col = city_col = None
    for i, h in enumerate(headers):
        if header_matches(h, COMPANY_PATTERNS):
            company_col = i
        elif header_matches(h, CITY_PATTERNS):
            city_col = i

    if company_col is None:
        print(f"Warning: No company column in sheet '{ws.title}'.", file=sys.stderr)
        return []

    data_cols = [
        (i, h) for i, h in enumerate(headers)
        if i not in (company_col, city_col) and h and not header_matches(h, ID_PATTERNS)
    ]

    count_cols, index_cols, untyped_cols = [], [], []
    for col_idx, col_header in data_cols:
        col_type = detect_column_type(col_header)
        if col_type == "count":
            count_cols.append((col_idx, col_header, strip_type_patterns(col_header, COUNT_PATTERNS)))
        elif col_type == "index":
            index_cols.append((col_idx, col_header, strip_type_patterns(col_header, INDEX_PATTERNS)))
        else:
            untyped_cols.append((col_idx, col_header))

    # Pair count/index columns by matching category name
    categories = []
    used_counts, used_indexes = set(), set()
    for ci, (c_idx, c_header, c_cat) in enumerate(count_cols):
        for ii, (i_idx, i_header, i_cat) in enumerate(index_cols):
            if ii in used_indexes:
                continue
            if c_cat and i_cat and c_cat == i_cat:
                categories.append({"name": c_cat.replace(" ", "_"), "count_col": c_idx, "index_col": i_idx})
                used_counts.add(ci)
                used_indexes.add(ii)
                break

    for ci, (c_idx, c_header, _) in enumerate(count_cols):
        if ci not in used_counts:
            categories.append({"name": c_header.lower().replace(" ", "_"), "value_col": c_idx, "field": "count"})
    for ii, (i_idx, i_header, _) in enumerate(index_cols):
        if ii not in used_indexes:
            categories.append({"name": i_header.lower().replace(" ", "_"), "value_col": i_idx})
    for col_idx, col_header in untyped_cols:
        categories.append({"name": col_header.lower().replace(" ", "_"), "value_col": col_idx})

    companies = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if company_col >= len(row) or not row[company_col]:
            continue
        company_name = str(row[company_col]).strip()
        city_name = str(row[city_col]).strip() if city_col is not None and city_col < len(row) and row[city_col] else ""
        entry = {"company": company_name, "city": city_name, "categories": {}}

        for cat in categories:
            cat_name = cat["name"]
            if "count_col" in cat and "index_col" in cat:
                count_val = index_val = None
                try:
                    count_val = int(parse_numeric_cell(row[cat["count_col"]]))
                except (ValueError, TypeError, IndexError):
                    pass
                try:
                    index_val = parse_numeric_cell(row[cat["index_col"]])
                except (ValueError, TypeError, IndexError):
                    pass
                if count_val is None and index_val is None:
                    continue
                entry["categories"][cat_name] = {"count": count_val, "index": index_val}
            elif "value_col" in cat:
                try:
                    val = parse_numeric_cell(row[cat["value_col"]])
                except (ValueError, TypeError, IndexError):
                    continue
                field = cat.get("field", "index")
                entry["categories"][cat_name] = {field: int(val) if field == "count" else val}

        companies.append(entry)

    return companies


def main():
    parser = argparse.ArgumentParser(description="Convert salary Excel data to JSON")
    parser.add_argument("excel_file", help="Path to the Excel file with salary data")
    parser.add_argument("--output", default=None, help="Output JSON path (default: salary_data.json in repo root)")
    parser.add_argument("--source", default=None, help="Data source name")
    parser.add_argument("--baseline", type=float, default=100, help="Baseline value for index (default: 100)")
    parser.add_argument("--baseline-desc", default=None, help="Description of what baseline means")
    args = parser.parse_args()

    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    if openpyxl is None:
        print("Error: openpyxl is required. Install: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output) if args.output else Path(__file__).parent.parent / "salary_data.json"

    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    all_companies = []
    for sheet_name in wb.sheetnames:
        print(f"  Parsing sheet: {sheet_name}")
        all_companies.extend(parse_sheet(wb[sheet_name], sheet_label=sheet_name))
    wb.close()

    if not all_companies:
        print("Error: No data could be parsed. Ensure the Excel file has a Company/Firma column.", file=sys.stderr)
        sys.exit(1)

    output = {
        "metadata": {
            "source": args.source or excel_path.stem,
            "index_baseline": args.baseline,
            "index_label": "Index",
            "baseline_description": args.baseline_desc or f"Index {args.baseline} = baseline",
        },
        "companies": all_companies,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nDone! Wrote {len(all_companies)} company entries to {output_path}")


if __name__ == "__main__":
    main()
